from enum import Enum

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime
from datetime import timedelta
from scipy import integrate
import numpy


class DataType(Enum):
    Level = 1
    Velocity = 2
    Temperature = 3


# Minimum Velocity in feet/sec
valid_velocity_min = float(input("Enter valid minimum value for velocity in feet/sec: "))

# Maximum Velocity in feet/sec
valid_velocity_max = float(input("Enter valid maximum value for velocity in feet/sec: "))

# Minimum Depth in inches
valid_level_min = float(input("Enter valid min value for level in inches: "))

# Minimum Depth in inches
valid_level_max = float(input("Enter valid max value for level in inches: "))

# Pipe radius in inches
pipe_radius = float(input("Enter the internal pipe diameter in inches: ")) / 2.0

# Minimum percentage of good data needed
minimum_percentage = float(input("Enter the minimum percentage of good data required to form average datasets: "))


class Entry:

    def __init__(self, timestamp: datetime, level, velocity, temperature):
        self.timestamp = timestamp
        self.level = level  # Inches
        self.velocity = velocity  # Feet/sec
        self.temperature = temperature
        self.flow = None  # Gallons/min

        if (valid_velocity_min < float(self.velocity) < valid_velocity_max) and (valid_level_min < float(self.level) < valid_level_max):
            self.is_valid = True
        else:
            self.is_valid = False


def get_nth_occurrence(string, sub_str, occurrence):
    val = -1
    for i in range(0, occurrence):
        val = string.find(sub_str, val + 1)
    return val


# This function takes the raw string and cleans it up between different data types.
# For example, it would clean up all of the Level(ft) data into "Date\nValue\n" format, returning a list of lines like
# that.
def get_single_data(raw_string):
    temp_string = raw_string  # Copy String
    last_useful_pos = get_nth_occurrence(temp_string[temp_string.rfind("#", ):], "\n", 2) + temp_string.rfind("#", )  #

    temp_string = temp_string[temp_string.find("#"):last_useful_pos]

    temp_string = temp_string.replace("#", "")

    all_lines = temp_string.splitlines(False)

    return all_lines


def parse_datetime(string):
    if len(string) > 11:
        # print("Long", datetime.strptime(string, '%Y-%m-%d %H:%M:%S'))
        return datetime.strptime(string, '%Y-%m-%d %H:%M:%S')
    else:
        # print("Short", datetime.strptime(string, '%Y-%m-%d'))
        return datetime.strptime(string, '%Y-%m-%d')


def get_all_entries(raw_string):
    level_start = raw_string.find("Level")
    velocity_start = raw_string.find("Velocity")
    temperature_start = raw_string.find("Temperature")

    entries = [get_single_data(raw_string[level_start:velocity_start]),
               get_single_data(raw_string[velocity_start:temperature_start]),
               get_single_data(raw_string[temperature_start:])]

    return entries


print("Loading Dat files into memory...")

file_num = 1
raw_strings = []

while os.path.isfile(str(file_num) + ".Dat"):
    f = open(str(file_num) + ".Dat")
    print(str(file_num) + ".Dat")
    raw_strings.append(f.read())
    f.close()
    file_num += 1

print(len(raw_strings), "files loaded.")

# print(raw_strings)

entry_list = []

for file in raw_strings:
    entry_string = get_all_entries(file)
    level_data = entry_string[0]
    velocity_data = entry_string[1]
    temperature_data = entry_string[2]

    # print(level_data)

    # print(len(level_data))
    # print(len(velocity_data))
    # print(len(temperature_data))



    # This assumes that the level, velocity, and temperature sections have the same number of entries, and that each entry in order has the same date as other entries in that position.

    for i in range(0, len(level_data), 2):
        entry_list.append(
            Entry(parse_datetime(level_data[i]), level_data[i + 1], velocity_data[i + 1], temperature_data[i + 1]))
    print("Finished parsing file.")

print("Creating Excel File...")

wb = Workbook()
ws = wb.active
ws.title = "Main"
minws = wb.create_sheet(title="15 Minute")
hourws = wb.create_sheet(title="1 Hour")
dayws = wb.create_sheet(title="1 Day")
minchartws = wb.create_sheet(title="15 Minute Chart")
hourchartws = wb.create_sheet(title="Hour Chart")
daychartws = wb.create_sheet(title="Day Chart")

ws['A1'] = "Date and Time"
ws['B1'] = "Level (inches)"
ws['C1'] = "Level (ft)"
ws['D1'] = "Velocity (ft/sec)"
ws['E1'] = "Velocity (ft/min)"
ws['F1'] = "Temperature (°F)"
ws['G1'] = "Flow (USG/min)"

minws['A1'] = "Date and Time Start"
minws['B1'] = "Average 15 minute data (USG/min)"
minws['C1'] = "Avg 15 min flow (USG/15 min)"
minws['D1'] = "Percentage of Good Data"

hourws['A1'] = "Date and Time Start"
hourws['B1'] = "Average 1 hour data (USG/min)"
hourws['C1'] = "Avg hour flow (USG/hour)"
hourws['D1'] = "Percentage of Good Data"

dayws['A1'] = "Date and Time Start"
dayws['B1'] = "Average 1 day data (USG/min)"
dayws['C1'] = "Avg day flow (USG/day)"
dayws['D1'] = "Percentage of Good Data"

last_min_row = None
last_hour_row = None
last_day_row = None


y = lambda x: (numpy.sqrt(numpy.square(pipe_radius) - numpy.square(x - pipe_radius)))

# print(integrate.quad(y, 0, 16))

last_print_percentage = 0


# Integrate to find flow values.

for i, entry in enumerate(entry_list):
    ws['A' + str(i + 2)] = entry.timestamp
    ws['B' + str(i + 2)] = float(entry.level)
    ws['C' + str(i + 2)] = "=B" + str(i + 2) + "/12"
    ws['D' + str(i + 2)] = float(entry.velocity)
    ws['E' + str(i + 2)] = "=D" + str(i + 2) + "*60"
    ws['F' + str(i + 2)] = round(float(entry.temperature), 2)
    if entry.is_valid:
        # Calculate square inches. This is cross section of pipe from bottom of pipe up to depth level.
        integral = integrate.quad(y, 0, float(entry.level))[0]

        # We multiply by velocity (converted to inches/sec) * the cross section (in inches^2) * 0.004329 (inches^3 to gallons) * 2 (The integral is only half of what it should be) * 60.0 (convert secs to mins) to get gallons/min flow through the pipe.
        entry.flow = round(integral * (float(entry.velocity) * 12.0) * 0.004329 * 2 * 60.0, 3)
        ws['G' + str(i + 2)] = entry.flow  # Gallons / min
    else:
        ws['G' + str(i + 2)] = '-'

    if (int(i / len(entry_list) * 100) % 10 == 0) and (int(i / len(entry_list) * 100) > last_print_percentage):
        print(str(round(i / len(entry_list) * 100, 2)) + "%")
        last_print_percentage = round(i / len(entry_list) * 100, 0)

first_time = entry_list[0].timestamp
last_time = entry_list[-1].timestamp

print("100.0%\nFinding 15 minute averages.")
current_time = first_time
current_ending_time = current_time + timedelta(minutes=15)
index = 2
start_ref_index = 2
end_ref_index = 2
current_entry_index = 0
keep_going = True
while keep_going:
    good_data = 0
    bad_data = 0
    for i, entry in enumerate(entry_list[current_entry_index:]):
        # print(current_entry_index + i)
        if current_time <= entry.timestamp < current_ending_time:
            if entry.is_valid:
                good_data += 1
            else:
                bad_data += 1
            end_ref_index = start_ref_index + i
        else:
            current_entry_index += i
            break
    # print("end index:", current_entry_index)

    minws['A' + str(index)] = current_time
    if good_data + bad_data > 0:
        percentage_good_data = round((good_data/(good_data + bad_data) * 100), 2)
    else:
        percentage_good_data = 0
    if good_data > 0 and percentage_good_data >= minimum_percentage:
        minws['B' + str(index)] = "=AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ")"
        minws['C' + str(index)] = "=(AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ") * 15)"
    else:
        minws['B' + str(index)] = "-"
        minws['C' + str(index)] = "-"
    minws['D' + str(index)] = percentage_good_data

    last_min_row = index

    start_ref_index = end_ref_index + 1
    index += 1
    current_time = current_ending_time
    current_ending_time = current_time + timedelta(minutes=15)
    if current_ending_time > last_time:
        keep_going = False

print("100.0%\nFinding 1 hour averages.")
current_time = first_time
current_ending_time = current_time + timedelta(hours=1)
index = 2
start_ref_index = 2
end_ref_index = 2
current_entry_index = 0
keep_going = True
while keep_going:
    good_data = 0
    bad_data = 0
    for i, entry in enumerate(entry_list[current_entry_index:]):
        # print(current_entry_index + i)
        if current_time <= entry.timestamp < current_ending_time:
            if entry.is_valid:
                good_data += 1
            else:
                bad_data += 1
            end_ref_index = start_ref_index + i
        else:
            current_entry_index += i
            break
    # print("end index:", current_entry_index)

    hourws['A' + str(index)] = current_time
    percentage_good_data = round((good_data/(good_data + bad_data)) * 100, 2)
    if good_data > 0 and percentage_good_data >= minimum_percentage:
        hourws['B' + str(index)] = "=AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ")"
        hourws['C' + str(index)] = "=(AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ") * 60)"
    else:
        hourws['B' + str(index)] = "-"
        hourws['C' + str(index)] = "-"
    hourws['D' + str(index)] = percentage_good_data

    last_hour_row = index

    start_ref_index = end_ref_index + 1
    index += 1
    current_time = current_ending_time
    current_ending_time = current_time + timedelta(hours=1)
    if current_ending_time > last_time:
        keep_going = False


print("Finding 1 day averages.")
current_time = first_time
current_ending_time = current_time + timedelta(days=1)
index = 2
start_ref_index = 2
end_ref_index = 2
current_entry_index = 0
keep_going = True
while keep_going:
    good_data = 0
    bad_data = 0
    for i, entry in enumerate(entry_list[current_entry_index:]):
        # print(current_entry_index + i)
        if current_time <= entry.timestamp < current_ending_time:
            if entry.is_valid:
                good_data += 1
            else:
                bad_data += 1
            end_ref_index = start_ref_index + i
        else:
            current_entry_index += i
            break
    # print("end index:", current_entry_index)

    dayws['A' + str(index)] = current_time
    percentage_good_data = round((good_data/(good_data + bad_data) * 100), 2)
    if good_data > 0 and percentage_good_data >= minimum_percentage:
        dayws['B' + str(index)] = "=AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ")"
        dayws['C' + str(index)] = "=(AVERAGE(Main!G" + str(start_ref_index) + ":G" + str(end_ref_index) + ") * 1440)"
    else:
        dayws['B' + str(index)] = "-"
        dayws['C' + str(index)] = "-"
    dayws['D' + str(index)] = percentage_good_data

    last_day_row = index

    start_ref_index = end_ref_index + 1
    index += 1
    current_time = current_ending_time
    current_ending_time = current_time + timedelta(days=1)
    if current_ending_time > last_time:
        keep_going = False

print("Creating Charts")

avg15chart = BarChart()
avg15chart.type = "col"
avg15chart.style = "1"
avg15chart.grouping = "clustered"
avg15chart.title = "Average 15-minute Flow Data from Flow Monitoring"
avg15chart.y_axis.title = "Flow (USG/min)"
avg15chart.x_axis.title = "Date"
data = Reference(minws, min_col=2, max_col=2, min_row=1, max_row=last_min_row)
cats = Reference(minws, min_col=1, max_col=1, min_row=2, max_row=last_min_row)
avg15chart.add_data(data, titles_from_data=True)
avg15chart.set_categories(cats)
minchartws.add_chart(avg15chart, "A1")

avghourchart = BarChart()
avghourchart.type = "col"
avghourchart.style = "1"
avghourchart.grouping = "clustered"
avghourchart.title = "Average Hour Flow Data from Flow Monitoring"
avghourchart.y_axis.title = "Flow (USG/min)"
avghourchart.x_axis.title = "Date"
data = Reference(hourws, min_col=2, max_col=2, min_row=1, max_row=last_hour_row)
cats = Reference(hourws, min_col=1, max_col=1, min_row=2, max_row=last_hour_row)
avghourchart.add_data(data, titles_from_data=True)
avghourchart.set_categories(cats)
hourchartws.add_chart(avghourchart, "A1")

avgdaychart = BarChart()
avgdaychart.type = "col"
avgdaychart.style = "1"
avgdaychart.grouping = "clustered"
avgdaychart.title = "Average Day Flow Data from Flow Monitoring"
avgdaychart.y_axis.title = "Flow (USG/min)"
avgdaychart.x_axis.title = "Date"
data = Reference(dayws, min_col=2, max_col=2, min_row=1, max_row=last_day_row)
cats = Reference(dayws, min_col=1, max_col=1, min_row=2, max_row=last_day_row)
avgdaychart.add_data(data, titles_from_data=True)
avgdaychart.set_categories(cats)
daychartws.add_chart(avgdaychart, "A1")

# Second chart.
avgdaychart2 = BarChart()
avgdaychart2.type = "col"
avgdaychart2.style = "1"
avgdaychart2.grouping = "clustered"
avgdaychart2.title = "Average Day Flow Data from Flow Monitoring"
avgdaychart2.y_axis.title = "Flow (USG/day)"
avgdaychart2.x_axis.title = "Date"
data = Reference(dayws, min_col=3, max_col=3, min_row=1, max_row=last_day_row)
cats = Reference(dayws, min_col=1, max_col=1, min_row=2, max_row=last_day_row)
avgdaychart2.add_data(data, titles_from_data=True)
avgdaychart2.set_categories(cats)
daychartws.add_chart(avgdaychart2, "A16")


print("Saving excel file...")

wb.save("output.xlsx")

print("Finished Creating Excel File.")
